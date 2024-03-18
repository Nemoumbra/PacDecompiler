import collections
import math
import random
from argparse import ArgumentParser

from Core.PAC.pac_file import (
    PAC_file, Padding_bytes, PAC_instruction,
)
from Core.PAC.pac_parser import (
    PAC_parser
)

from Core.PAC.pac_dumper import (
    PAC_DisasmSettings, PAC_dumper
)

from Utils.utils import (
    load_file_by_path
)

from Core.PAC.instruction_set_reader import (
    InstructionSetReader
)

from typing import Dict, Set
from pathlib import Path
from xlsxwriter import Workbook
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Border, Side


import decompiler_tests


def parse_args():
    parser = ArgumentParser('Mode to run')
    parser.add_argument(
        '--mode',
        choices=('decomp_tests', 'version_tracking', "decompile_dir", "scripts")
    )
    return parser.parse_args()


def study_pac_data_types(directory: Path, instruction_set: Path):
    instr_set_reader = InstructionSetReader()
    pac_parser = PAC_parser()
    instr_set_reader.read_instruction_set(str(instruction_set))
    pac_parser.setTemplates(instr_set_reader.PAC_instruction_templates)

    # pac_parser.cmd_inxJmp_signature = 0x25002D00  # P2
    pac_parser.cmd_inxJmp_signature = 0x25002f00  # P3
    files = directory.glob("*.pac")
    for path in files:
        if not path.is_file():
            continue
        try:
            file = PAC_file()
            full_path = directory / path.name
            file.initialize_by_raw_data(load_file_by_path(str(full_path)))

            pac_parser.reset(file)
            pac_parser.parse()
            print(f"{path.name} parsed successfully!")

        except Exception as e:
            print(e)


def add_file_stats_to_heatmap(pac_file: PAC_file, mapping: Dict[str, int], worksheet):
    last_was_instruction = False
    last_index = -1
    for entity in pac_file.entities.values():
        entity_type = type(entity)
        if entity_type is Padding_bytes:
            continue

        if entity_type is PAC_instruction:
            entity_type: PAC_instruction
            if last_was_instruction:
                this_index = mapping[entity.name]

                cell = worksheet.cell(last_index, this_index)
                new_value = cell.value + 1 if cell.value is not None else 1
                # cell.border = make_border_for_cell()
                worksheet.cell(last_index, this_index, new_value)

                last_index = this_index
            else:
                last_was_instruction = True
                last_index = mapping[entity.name]
        else:
            last_was_instruction = False


def excel_heatmap(directory: Path, instruction_set: Path, xml_path: Path):
    wb = load_workbook(xml_path)
    worksheet = wb.worksheets[0]

    instr_set_reader = InstructionSetReader()
    instr_set_reader.read_instruction_set(str(instruction_set))
    pac_parser = PAC_parser()
    # pac_parser.cmd_inxJmp_signature = 0x25002D00  # P2
    pac_parser.cmd_inxJmp_signature = 0x25002f00  # P3
    pac_parser.setTemplates(instr_set_reader.PAC_instruction_templates)
    mapping = {template.name: i+2 for i, template in enumerate(instr_set_reader.PAC_instruction_templates.values())}

    files = directory.glob("*.pac")
    for path in files:
        if not path.is_file():
            continue
        try:
            file = PAC_file()
            full_path = directory / path.name
            file.initialize_by_raw_data(load_file_by_path(str(full_path)))
            pac_parser.reset(file)
            pac_parser.parse()
            print(f"{path.name} parsed successfully...")

            add_file_stats_to_heatmap(file, mapping, worksheet)
            print(f"{path.name} pushed to the worksheet!")
        except Exception as e:
            print(e)

    wb.save(str(xml_path.with_suffix("")) + "_edited.xlsx")
    print("Done!")
    exit()


def get_consecutive_instructions(pac_file: PAC_file):
    last_was_instruction = False
    last_offset = -1

    for offset, entity in pac_file.entities.items():
        entity_type = type(entity)
        if entity_type is Padding_bytes:
            continue

        if entity_type is PAC_instruction:
            entity_type: PAC_instruction
            if last_was_instruction:
                # Let's yield the pair
                yield last_offset, offset

                last_offset = offset
            else:
                last_was_instruction = True
        else:
            last_was_instruction = False
    # finish the generator
    return


def sample_pair_data(pac_file: PAC_file):
    offset_pairs = list(get_consecutive_instructions(pac_file))
    count = math.ceil(math.sqrt(len(offset_pairs)))
    return collections.Counter(random.choices(offset_pairs, k=count))


def better_excel_heatmap(directory: Path, instruction_set: Path):
    instr_set_reader = InstructionSetReader()
    instr_set_reader.read_instruction_set(str(instruction_set))
    pac_parser = PAC_parser()
    # pac_parser.cmd_inxJmp_signature = 0x25002D00  # P2
    pac_parser.cmd_inxJmp_signature = 0x25002f00  # P3
    pac_parser.setTemplates(instr_set_reader.PAC_instruction_templates)

    files = directory.glob("*.pac")
    for path in files:
        if not path.is_file():
            continue
        try:
            file = PAC_file()
            full_path = directory / path.name
            file.initialize_by_raw_data(load_file_by_path(str(full_path)))
            pac_parser.reset(file)
            pac_parser.parse()
            print(f"{path.name} parsed successfully...")

            data = sample_pair_data(file)
        except Exception as e:
            print(e)

    print("Done!")
    exit()


class PAC_test_base:
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int):
        self.directory = directory

        self.instr_set_reader = InstructionSetReader()
        self._pac_parser = PAC_parser()
        self.instr_set_reader.read_instruction_set(str(instruction_set))
        self._pac_parser.setTemplates(self.instr_set_reader.PAC_instruction_templates)

        self._pac_parser.cmd_inxJmp_signature = cmd_inxJmp

    def setup(self, instr_set_reader: InstructionSetReader):
        pass

    def fini(self):
        pass

    def test(self, file: PAC_file, path: Path):
        raise NotImplementedError

    def __call__(self):
        files = self.directory.glob("*.pac")
        for path in files:
            if not path.is_file():
                continue
            try:
                file = PAC_file()
                full_path = self.directory / path.name
                file.name = path.name
                file.initialize_by_raw_data(load_file_by_path(str(full_path)))

                self._pac_parser.reset(file)
                self._pac_parser.parse()
                print(f"{path.name} parsed successfully!")

                res = self.test(file, path)
                if res is not None and not res:
                    print("Stopping iteration (received command to halt)")
                    break

            except Exception as e:
                print(e)

        self.fini()
        exit()


class PAC_networkErrorInvokeTester(PAC_test_base):
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int):
        super().__init__(directory, instruction_set, cmd_inxJmp)
        self.counterexamples = {}

    def test(self, file: PAC_file, path: Path):
        # 0x251c2c00 is 'networkErrorInvoke'
        all_networkErrorInvoke = file.getInstructions(0x25030200)
        for location in all_networkErrorInvoke:
            offset, next_object = file.get_entity_by_offset(location + 16)
            if type(next_object) is not PAC_instruction:
                self.counterexamples[f"{file.name}"] = offset
                return
            next_object: PAC_instruction
            if next_object.name != "cmd_end":
                self.counterexamples[f"{file.name}"] = offset
                return

    def fini(self):
        if not self.counterexamples:
            print("\nThe hypothesis is correct!")
            return
        print("\nCounterexamples found!")
        for filename, offset in self.counterexamples.items():
            print(f"{filename}: offset 0x{offset:X}")


class PAC_disassemble_in_dir_tester(PAC_test_base):
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int, settings: PAC_DisasmSettings):
        super().__init__(directory, instruction_set, cmd_inxJmp)
        self.settings = settings

    def test(self, file: PAC_file, path: Path):
        dumper = PAC_dumper()
        dumper.reset(file, self.settings)
        dumper.disassemble_to_file(self.directory / (path.with_suffix(".txt")))


class PAC_find_inconsistencies_in_dir_tester(PAC_test_base):
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int, where_to: Path):
        super().__init__(directory, instruction_set, cmd_inxJmp)
        self.unknown_signatures = set()
        self.potential_incorrect_signatures_map: Dict[str, Set[int]] = {}

        # Don't forget to close that in 'fini'
        self.output = open(where_to, "w", encoding="utf-8")

    def test(self, file: PAC_file, path: Path):
        self.unknown_signatures = self.unknown_signatures.union(file.unknown_instructions.keys())

        previous_section = -1
        if file.unknown_instructions:
            print("Unknown instructions:")
            self.output.write("Unknown instructions:\n")
        for signature in sorted(file.unknown_instructions.keys()):
            section = (signature // 65536) % 256
            if section != previous_section:
                previous_section = section
                print(f"Section 0x{section:X}")
                self.output.write(f"Section 0x{section:X}\n")
            print(f"{signature:X}")
            self.output.write(f"{signature:X}\n")

        # if file.left_out_PAC_arguments:
        #     print("Potential left out PAC args:")
        #     output.write("Potential left out PAC args:\n")
        for arg in file.left_out_PAC_arguments.values():
            if path.name not in self.potential_incorrect_signatures_map:
                self.potential_incorrect_signatures_map[path.name] = set()
            self.potential_incorrect_signatures_map[path.name].add(arg.supposed_signature)

    def fini(self):
        print("\nEverything parsed! To sum it up...")
        self.output.write("\nEverything parsed! To sum it up...\n")

        if self.unknown_signatures:
            print("Found unknown instructions:")
            self.output.write("Found unknown instructions:\n")
            previous_section = -1
            for signature in sorted(list(self.unknown_signatures)):
                section = (signature // 65536) % 256
                if section != previous_section:
                    previous_section = section
                    print(f"Section 0x{section:X}")
                    self.output.write(f"Section 0x{section:X}\n")
                print(f"{signature:X}")
                self.output.write(f"{signature:X}\n")
        else:
            print("No unknown instructions found!")
            self.output.write("No unknown instructions found!\n")

        if self.potential_incorrect_signatures_map:
            print("Potential incorrect signatures:")
            self.output.write("Potential incorrect signatures:\n")
            potential_incorrect_signatures = "\n".join(
                (f"{signature:X}" for signature in set().union(*self.potential_incorrect_signatures_map.values()))
            )
            print(potential_incorrect_signatures)
            self.output.write(potential_incorrect_signatures + "\n")
            print("Found in files:")
            self.output.write("Found in files:\n")
            found_in_files = "\n".join(self.potential_incorrect_signatures_map.keys())
            print(found_in_files)
            self.output.write(found_in_files + "\n")
        else:
            print("No incorrect signatures found!")
            self.output.write("No incorrect signatures found!\n")

        # Close the output:
        self.output.close()


class PAC_instructions_usage_to_excel_tester(PAC_test_base):
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int, xml_path: Path):
        super().__init__(directory, instruction_set, cmd_inxJmp)

        self.workbook = Workbook(xml_path)
        self.worksheet = self.workbook.add_worksheet("by Nemoumbra")
        self.signature_to_column: Dict[int, int] = {}

        self.row: int = 0
        self.column: int = 0

        # I don't want to enumerate here
        for signature, template in self.instr_set_reader.PAC_instruction_templates.items():
            self.row += 1
            self.signature_to_column[signature] = self.row
            self.worksheet.write(self.row, self.column, f"{signature:X} ({template.name})")
        self.column = 5

    def test(self, file: PAC_file, path: Path):
        self.column += 1
        self.row = 0
        self.worksheet.write(self.row, self.column, path.name)
        for signature, places in file.instructions.items():
            count = len(places)
            row = self.signature_to_column[signature]
            self.worksheet.write(row, self.column, count)
        print(f"{path.name} pushed to the worksheet!")

    def fini(self):
        self.workbook.close()


class PAC_get_resource_strings_tester(PAC_test_base):
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int, where_to: Path):
        super().__init__(directory, instruction_set, cmd_inxJmp)
        self.pac_strings: Set[str] = set()
        self.where_to = where_to

    def test(self, file: PAC_file, path: Path):
        all_readArcFile = file.getInstructions(0x25090900)
        for location, instruction in all_readArcFile.items():
            arguments = instruction.ordered_PAC_params
            filename = arguments[1][1]
            self.pac_strings.add(filename)
        pass

    def fini(self):
        with open(self.where_to, mode="w") as output:
            output.write("\n".join(self.pac_strings))


class PAC_get_0x1_arg_instr_tester(PAC_test_base):
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int):
        super().__init__(directory, instruction_set, cmd_inxJmp)
        self._0x1_instructions: Set[int] = set()

    def test(self, file: PAC_file, path: Path):
        for instruction in file.ordered_instructions.values():
            _0x1_args = instruction.get_used_0x1_values()
            if _0x1_args:
                self._0x1_instructions.add(instruction.signature)

    def fini(self):
        if not self._0x1_instructions:
            print("No instructions that may accept 0x1 args found!")
            return

        print("\nFound instructions that can accept 0x1 args:")
        print(
            "\n".join(
                [f"{i:X} - {self.instr_set_reader.PAC_signature_to_name[i]}" for i in sorted(self._0x1_instructions)]
            )
        )


class PAC_get_4byte_value_instr_tester(PAC_test_base):
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int):
        super().__init__(directory, instruction_set, cmd_inxJmp)
        self._4_byte_value_instructions: Set[int] = set()

    def test(self, file: PAC_file, path: Path):
        for instruction in file.ordered_instructions.values():
            _4_byte_value_args = instruction.get_used_4_byte_values()
            if _4_byte_value_args:
                self._4_byte_value_instructions.add(instruction.signature)

    def fini(self):
        if not self._4_byte_value_instructions:
            print("No instructions that may accept 4 byte values found!")
            return

        print("\nFound instructions that can accept 4 byte values:")
        print(
            "\n".join(
                [
                    f"{i:X} - {self.instr_set_reader.PAC_signature_to_name[i]}"
                    for i in sorted(self._4_byte_value_instructions)
                ]
            )
        )


class PAC_cmd_stk_checker_tester(PAC_test_base):
    def __init__(self, directory: Path, instruction_set: Path, cmd_inxJmp: int):
        super().__init__(directory, instruction_set, cmd_inxJmp)

        cmd_stkDec = 0x25003000
        cmd_stkClr = 0x25003100
        self.signatures = (cmd_stkDec, cmd_stkClr)

        self.counterexamples: Dict[int, Dict[str, int]] = {
            cmd_stkDec: {}, cmd_stkClr: {}
        }

    def test(self, file: PAC_file, path: Path):
        for signature in self.signatures:
            instructions = file.getInstructions(signature)

            # We break from this loop to move onto the next signature
            for location in instructions:
                offset, next_object = file.get_entity_by_offset(location + 4)
                if type(next_object) is not PAC_instruction:
                    self.counterexamples[signature][file.name] = offset
                    break
                next_object: PAC_instruction
                if next_object.name not in ("cmd_end", "cmd_jmp"):
                    self.counterexamples[signature][file.name] = offset
                    break

    def fini(self):
        bad = self.counterexamples[self.signatures[0]] and self.counterexamples[self.signatures[1]]
        if not bad:
            print("\nThe hypothesis is correct!")
            return
        print("\nCounterexamples found!")
        for signature, mapping in self.counterexamples.items():
            for filename, offset in mapping.items():
                print(f"{filename}: offset 0x{offset:X}")


def main():
    print("Program started!")
    input("Press enter")

    # test_func = PAC_networkErrorInvokeTester(
    #     Path(input()),
    #     Path(input()),
    #     0x25002f00,  # P3
    #     # 0x25002D00,  # P1/2
    # )
    # test_func()

    # cmd_stk_check_func = PAC_cmd_stk_checker_tester(
    #     Path(input()),
    #     Path(input()),
    #     0x25002f00,  # P3
    #     # 0x25002D00,  # P1/2
    # )
    # cmd_stk_check_func()

    # instructions_usage_to_excel_func = PAC_instructions_usage_to_excel_tester(
    #     Path(input()),
    #     Path(input()),
    #     0x25002f00,  # P3
    #     # 0x25002D00,  # P1/2
    #     Path(input())
    # )
    # instructions_usage_to_excel_func()

    # excel_heatmap(
    #     Path(input()),
    #     Path(input()),
    #     Path(input()),
    # )

    # better_excel_heatmap(
    #     Path(input()),
    #     Path(input())
    # )

    # get_strings_from_pacs(
    #     Path(input()),
    #     Path(input()),
    #     Path(input())
    # )

    # get_resource_strings_func = PAC_get_resource_strings_tester(
    #     Path(input()),
    #     Path(input()),
    #     0x25002f00,  # P3
    #     # 0x25002D00,  # P1/2
    #     Path(input())
    # )
    # get_resource_strings_func()

    # get_0x1_instructions_func = PAC_get_0x1_arg_instr_tester(
    #     Path(input()),
    #     Path(input()),
    #     0x25002f00,  # P3
    #     # 0x25002D00,  # P1/2
    # )
    # get_0x1_instructions_func()

    # get_4_byte_value_instructions(
    #     Path(input()),
    #     Path(input())
    # )

    # get_4_byte_value_instructions_func = PAC_get_4byte_value_instr_tester(
    #     Path(input()),
    #     Path(input()),
    #     0x25002f00,  # P3
    #     # 0x25002D00,  # P1/2
    # )
    # get_4_byte_value_instructions_func()

    # study_flags_and_globals(
    #     Path(input()),
    #     Path(input()),
    #     Path(input()),
    # )

    # find_file_inconsistencies_in_directory(
    #     Path(input()),
    #     Path(input()),
    #     Path(input())
    # )

    # find_PAC_inconsistencies_func = PAC_find_inconsistencies_in_dir_tester(
    #     Path(input()),
    #     Path(input()),
    #     0x25002f00,  # P3
    #     # 0x25002D00,  # P1/2
    #     Path(input()),
    # )
    # find_PAC_inconsistencies_func()

    # Contains the paths to useful files for the next call
    # disassemble_pacs_in_directory(
    #     Path(input()),
    #     Path(input())
    # )
    # exit()

    disasm_settings = PAC_DisasmSettings(
        omit_arg_names=True, skip_padding_bytes=False,
        decode_shift_jis=True, dump_failed_decodings=True
    )
    disasm_test_func = PAC_disassemble_in_dir_tester(
        Path(input()),
        Path(input()),
        0x25002f00,  # P3
        # 0x25002D00,  # P1/2
        disasm_settings
    )
    disasm_test_func()


if __name__ == '__main__':
    cmd_args = parse_args()
    print(f"{cmd_args.mode=}")
    if cmd_args.mode == "decomp_tests":
        decompiler_tests.run_tests()
    elif cmd_args.mode == "version_tracking":
        decompiler_tests.version_tracking_tests()
    elif cmd_args.mode == "decompile_dir":
        decompiler_tests.decompile_pacs_in_directory(
            Path(input()),
            Path(input())
        )
    elif cmd_args.mode == "scripts":
        main()
